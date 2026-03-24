"""
export_excel.py — Export orders to Excel (.xlsx) using openpyxl.
Supports optional start_date/end_date filtering. Streams as HttpResponse.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from orders.models import Order

HEADERS = ['Mã ĐH', 'Khách hàng', 'Ngày đặt', 'Ngày giao', 'Trạng thái', 'Tổng tiền (VND)']

# Header background: light blue
HEADER_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')


def export_orders_excel(start_date=None, end_date=None):
    """Build and return an Excel HttpResponse with filtered order data."""
    queryset = (
        Order.objects
        .select_related('customer')
        .order_by('order_date')
    )
    if start_date:
        queryset = queryset.filter(order_date__gte=start_date)
    if end_date:
        queryset = queryset.filter(order_date__lte=end_date)

    # Cap at 1000 rows to prevent timeout
    orders = queryset[:1000]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Don hang'

    # Header row styling
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, order in enumerate(orders, 2):
        ws.cell(row=row_idx, column=1, value=order.id)
        ws.cell(row=row_idx, column=2, value=order.customer.name)
        ws.cell(row=row_idx, column=3, value=str(order.order_date))
        ws.cell(row=row_idx, column=4, value=str(order.delivery_date) if order.delivery_date else '')
        ws.cell(row=row_idx, column=5, value=order.get_status_display())
        ws.cell(row=row_idx, column=6, value=float(order.total_amount))

    # Auto-size columns (approximate)
    col_widths = [10, 25, 12, 12, 18, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="don-hang.xlsx"'
    wb.save(response)
    return response
